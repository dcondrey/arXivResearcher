#!/usr/bin/env python3
"""
Comprehensive arXiv Paper Mining Script
========================================
Mines arXiv papers with rich metadata from multiple sources:
- Full paper metadata from arXiv API
- Citation data, venues, fields of study from Semantic Scholar
- Computed metrics (citation velocity, paper age, etc.)

Usage:
    uv run python mine_arxiv.py [options]

Examples:
    uv run python mine_arxiv.py --categories cs.AI,cs.CL --max-results 100
    uv run python mine_arxiv.py --start-date 2025-01-01 --end-date 2025-06-30
"""

import argparse
import csv
import time
import json
import urllib.request
import urllib.parse
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from collections import defaultdict
from typing import Optional, Dict, List, Any
import os
import re

# API endpoints
ARXIV_API_URL = "http://export.arxiv.org/api/query"
SEMANTIC_SCHOLAR_BATCH_API = "https://api.semanticscholar.org/graph/v1/paper/batch"

# Semantic Scholar fields to request
S2_FIELDS = [
    "citationCount",
    "influentialCitationCount",
    "referenceCount",
    "fieldsOfStudy",
    "publicationVenue",
    "openAccessPdf",
    "authors",
    "tldr",
    "publicationTypes",
    "externalIds",
    "s2FieldsOfStudy",
    "citationStyles"
]

# All arXiv CS categories
CS_CATEGORIES = [
    "cs.AI", "cs.AR", "cs.CC", "cs.CE", "cs.CG", "cs.CL", "cs.CR", "cs.CV",
    "cs.CY", "cs.DB", "cs.DC", "cs.DL", "cs.DM", "cs.DS", "cs.ET", "cs.FL",
    "cs.GL", "cs.GR", "cs.GT", "cs.HC", "cs.IR", "cs.IT", "cs.LG", "cs.LO",
    "cs.MA", "cs.MM", "cs.MS", "cs.NA", "cs.NE", "cs.NI", "cs.OH", "cs.OS",
    "cs.PF", "cs.PL", "cs.RO", "cs.SC", "cs.SD", "cs.SE", "cs.SI", "cs.SY"
]

# Additional categories for ML/Stats/Math
ADDITIONAL_CATEGORIES = [
    "stat.ML", "stat.TH", "stat.ME", "stat.AP",
    "math.OC", "math.ST", "math.NA",
    "eess.SP", "eess.IV", "eess.AS", "eess.SY",
    "q-bio.QM", "q-bio.NC",
    "quant-ph",
    "econ.EM", "econ.TH"
]

ALL_CATEGORIES = CS_CATEGORIES + ADDITIONAL_CATEGORIES


def parse_args():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="Mine arXiv papers with comprehensive metadata"
    )
    parser.add_argument(
        "--categories", type=str, default=None,
        help="Comma-separated arXiv categories (default: all CS categories)"
    )
    parser.add_argument(
        "--all-categories", action="store_true",
        help="Use all categories including stats, math, physics, etc."
    )
    parser.add_argument(
        "--max-results", type=int, default=500,
        help="Maximum results per category (default: 500)"
    )
    parser.add_argument(
        "--start-date", type=str, default=None,
        help="Start date YYYY-MM-DD (default: 12 months ago)"
    )
    parser.add_argument(
        "--end-date", type=str, default=None,
        help="End date YYYY-MM-DD (default: today)"
    )
    parser.add_argument(
        "--output-dir", type=str, default="arxiv_dataset_v2",
        help="Output directory (default: arxiv_dataset_v2)"
    )
    parser.add_argument(
        "--skip-semantic-scholar", action="store_true",
        help="Skip Semantic Scholar enrichment (faster)"
    )
    parser.add_argument(
        "--rate-limit", type=float, default=3.0,
        help="Seconds between arXiv API requests (default: 3.0)"
    )
    parser.add_argument(
        "--s2-rate-limit", type=float, default=1.0,
        help="Seconds between Semantic Scholar requests (default: 1.0)"
    )
    return parser.parse_args()


def fetch_arxiv_papers(category: str, start_date: str, end_date: str,
                       max_results: int = 500, rate_limit: float = 3.0) -> List[Dict]:
    """
    Fetch papers from arXiv API with FULL metadata.
    """
    papers = []
    batch_size = 100
    start_index = 0

    while start_index < max_results:
        query = f"cat:{category}"
        params = {
            "search_query": query,
            "start": start_index,
            "max_results": min(batch_size, max_results - start_index),
            "sortBy": "submittedDate",
            "sortOrder": "descending"
        }

        url = f"{ARXIV_API_URL}?{urllib.parse.urlencode(params)}"

        try:
            print(f"    Fetching batch {start_index}-{start_index + batch_size}...")

            req = urllib.request.Request(url, headers={"User-Agent": "arxiv-miner/2.0"})
            with urllib.request.urlopen(req, timeout=30) as response:
                xml_data = response.read().decode("utf-8")

            root = ET.fromstring(xml_data)
            ns = {
                "atom": "http://www.w3.org/2005/Atom",
                "arxiv": "http://arxiv.org/schemas/atom"
            }

            entries = root.findall("atom:entry", ns)

            if not entries:
                break

            for entry in entries:
                paper = extract_arxiv_entry(entry, ns, start_date, end_date, category)
                if paper:
                    papers.append(paper)

            start_index += batch_size
            time.sleep(rate_limit)

        except Exception as e:
            print(f"    Error: {e}")
            break

    return papers


def extract_arxiv_entry(entry, ns: dict, start_date: str, end_date: str,
                        default_category: str) -> Optional[Dict]:
    """Extract all available metadata from an arXiv entry."""

    # Paper ID
    id_elem = entry.find("atom:id", ns)
    paper_id = id_elem.text.split("/abs/")[-1] if id_elem is not None else ""

    # Extract version number
    version_match = re.search(r'v(\d+)$', paper_id)
    version = int(version_match.group(1)) if version_match else 1
    base_id = paper_id.split("v")[0] if "v" in paper_id else paper_id

    # Published date
    published = entry.find("atom:published", ns)
    pub_date = published.text[:10] if published is not None else ""
    pub_datetime = published.text if published is not None else ""

    # Filter by date range
    if pub_date:
        pub_date_clean = pub_date.replace("-", "")
        if pub_date_clean < start_date or pub_date_clean > end_date:
            return None

    # Updated date
    updated = entry.find("atom:updated", ns)
    updated_date = updated.text[:10] if updated is not None else ""
    updated_datetime = updated.text if updated is not None else ""

    # Title
    title_elem = entry.find("atom:title", ns)
    title = title_elem.text.strip().replace("\n", " ") if title_elem is not None else ""
    title = re.sub(r'\s+', ' ', title)  # Normalize whitespace

    # Authors with affiliations
    authors = []
    author_affiliations = []
    for author in entry.findall("atom:author", ns):
        name = author.find("atom:name", ns)
        if name is not None:
            authors.append(name.text)

        # Try to get affiliation (arxiv namespace)
        affil = author.find("arxiv:affiliation", ns)
        if affil is not None:
            author_affiliations.append(affil.text)

    # Primary category
    primary_cat = entry.find("arxiv:primary_category", ns)
    primary_category = primary_cat.get("term") if primary_cat is not None else default_category

    # All categories
    categories = []
    for cat in entry.findall("atom:category", ns):
        term = cat.get("term")
        if term:
            categories.append(term)

    # FULL abstract (not truncated!)
    summary = entry.find("atom:summary", ns)
    abstract = summary.text.strip().replace("\n", " ") if summary is not None else ""
    abstract = re.sub(r'\s+', ' ', abstract)

    # arXiv comment (e.g., "Accepted at NeurIPS 2025, 15 pages, 8 figures")
    comment = entry.find("arxiv:comment", ns)
    arxiv_comment = comment.text.strip() if comment is not None else ""

    # Journal reference
    journal_ref = entry.find("arxiv:journal_ref", ns)
    journal_reference = journal_ref.text.strip() if journal_ref is not None else ""

    # DOI
    doi_elem = entry.find("arxiv:doi", ns)
    doi = doi_elem.text if doi_elem is not None else ""

    # If no DOI in arxiv namespace, check links
    if not doi:
        for link in entry.findall("atom:link", ns):
            if link.get("title") == "doi":
                doi = link.get("href", "").replace("http://dx.doi.org/", "")
                break

    # Links
    pdf_url = ""
    abs_url = ""
    for link in entry.findall("atom:link", ns):
        rel = link.get("rel", "")
        link_type = link.get("type", "")
        href = link.get("href", "")

        if link_type == "application/pdf" or link.get("title") == "pdf":
            pdf_url = href
        elif rel == "alternate":
            abs_url = href

    if not abs_url:
        abs_url = f"https://arxiv.org/abs/{paper_id}"
    if not pdf_url:
        pdf_url = f"https://arxiv.org/pdf/{paper_id}.pdf"

    # Calculate paper age in days
    if pub_date:
        try:
            pub_dt = datetime.strptime(pub_date, "%Y-%m-%d")
            paper_age_days = (datetime.now() - pub_dt).days
        except ValueError:
            paper_age_days = None
    else:
        paper_age_days = None

    return {
        # Core identifiers
        "arxiv_id": paper_id,
        "arxiv_id_base": base_id,
        "version": version,

        # Core metadata
        "title": title,
        "authors": "; ".join(authors),
        "author_count": len(authors),
        "author_affiliations": "; ".join(author_affiliations) if author_affiliations else "",

        # Dates
        "published_date": pub_date,
        "published_datetime": pub_datetime,
        "updated_date": updated_date,
        "updated_datetime": updated_datetime,
        "paper_age_days": paper_age_days,

        # Categories
        "primary_category": primary_category,
        "all_categories": ", ".join(categories),
        "category_count": len(categories),

        # Content
        "abstract": abstract,
        "abstract_length": len(abstract),
        "arxiv_comment": arxiv_comment,
        "journal_reference": journal_reference,

        # Identifiers
        "doi": doi,

        # Links
        "arxiv_url": abs_url,
        "pdf_url": pdf_url,

        # Semantic Scholar fields (to be filled later)
        "citations": None,
        "influential_citations": None,
        "reference_count": None,
        "fields_of_study": "",
        "s2_fields_of_study": "",
        "publication_venue": "",
        "venue_type": "",
        "is_open_access": None,
        "open_access_url": "",
        "tldr": "",
        "publication_types": "",
        "s2_paper_id": "",
        "s2_url": "",

        # Computed metrics (to be filled after S2 enrichment)
        "citations_per_month": None,
        "citations_per_year": None,
    }


def enrich_with_semantic_scholar(papers: List[Dict], rate_limit: float = 1.0) -> List[Dict]:
    """
    Enrich papers with data from Semantic Scholar Graph API.
    Uses batch API for efficiency.
    """
    if not papers:
        return papers

    batch_size = 100  # S2 batch limit
    total = len(papers)

    for i in range(0, total, batch_size):
        batch = papers[i:i + batch_size]
        batch_ids = [f"ARXIV:{p['arxiv_id_base']}" for p in batch]

        print(f"    Enriching papers {i+1}-{min(i+batch_size, total)} of {total}...")

        try:
            request_data = json.dumps({"ids": batch_ids}).encode("utf-8")

            fields_param = ",".join(S2_FIELDS)
            url = f"{SEMANTIC_SCHOLAR_BATCH_API}?fields={fields_param}"

            req = urllib.request.Request(
                url,
                data=request_data,
                headers={
                    "User-Agent": "arxiv-miner/2.0",
                    "Content-Type": "application/json"
                },
                method="POST"
            )

            with urllib.request.urlopen(req, timeout=60) as response:
                results = json.loads(response.read().decode("utf-8"))

            # Map results back to papers
            for j, result in enumerate(results):
                paper = papers[i + j]

                if result is None:
                    # Paper not found in S2
                    paper["citations"] = 0
                    paper["influential_citations"] = 0
                    paper["reference_count"] = 0
                    continue

                # Citation data
                paper["citations"] = result.get("citationCount", 0)
                paper["influential_citations"] = result.get("influentialCitationCount", 0)
                paper["reference_count"] = result.get("referenceCount", 0)

                # Fields of study
                fos = result.get("fieldsOfStudy") or []
                paper["fields_of_study"] = ", ".join(fos)

                s2_fos = result.get("s2FieldsOfStudy") or []
                s2_fos_names = [f.get("category", "") for f in s2_fos if f.get("category")]
                paper["s2_fields_of_study"] = ", ".join(s2_fos_names)

                # Publication venue
                venue = result.get("publicationVenue") or {}
                paper["publication_venue"] = venue.get("name", "")
                paper["venue_type"] = venue.get("type", "")

                # Open access
                oa_pdf = result.get("openAccessPdf") or {}
                paper["is_open_access"] = oa_pdf.get("url") is not None
                paper["open_access_url"] = oa_pdf.get("url", "")

                # TL;DR
                tldr = result.get("tldr") or {}
                paper["tldr"] = tldr.get("text", "")

                # Publication types
                pub_types = result.get("publicationTypes") or []
                paper["publication_types"] = ", ".join(pub_types)

                # S2 identifiers
                paper["s2_paper_id"] = result.get("paperId", "")
                if paper["s2_paper_id"]:
                    paper["s2_url"] = f"https://www.semanticscholar.org/paper/{paper['s2_paper_id']}"

                # External IDs (for cross-referencing)
                ext_ids = result.get("externalIds") or {}
                if not paper["doi"] and ext_ids.get("DOI"):
                    paper["doi"] = ext_ids.get("DOI")

                # Author data with affiliations from S2
                s2_authors = result.get("authors") or []
                if s2_authors:
                    # S2 can have more detailed author info
                    s2_author_names = []
                    s2_affiliations = []
                    for auth in s2_authors:
                        if auth.get("name"):
                            s2_author_names.append(auth["name"])
                        # S2 sometimes has affiliations
                        if auth.get("affiliations"):
                            s2_affiliations.extend(auth["affiliations"])

                    # If we got affiliations from S2 and didn't have them before
                    if s2_affiliations and not paper["author_affiliations"]:
                        paper["author_affiliations"] = "; ".join(s2_affiliations[:10])

                # Compute citation velocity
                if paper["paper_age_days"] and paper["paper_age_days"] > 0:
                    months = max(paper["paper_age_days"] / 30.44, 1)  # At least 1 month
                    paper["citations_per_month"] = round(paper["citations"] / months, 3)
                    paper["citations_per_year"] = round(paper["citations"] / (months / 12), 3)

            time.sleep(rate_limit)

        except urllib.error.HTTPError as e:
            print(f"    S2 API error (HTTP {e.code}): {e.reason}")
            # Set defaults for this batch
            for j in range(len(batch)):
                papers[i + j]["citations"] = 0
                papers[i + j]["influential_citations"] = 0
                papers[i + j]["reference_count"] = 0
            time.sleep(rate_limit * 2)  # Back off on error

        except Exception as e:
            print(f"    S2 enrichment error: {e}")
            for j in range(len(batch)):
                papers[i + j]["citations"] = 0
                papers[i + j]["influential_citations"] = 0
                papers[i + j]["reference_count"] = 0

    return papers


def save_to_csv(papers_by_category: Dict[str, List[Dict]], output_dir: str):
    """Save papers to CSV files."""
    os.makedirs(output_dir, exist_ok=True)

    # Combine all papers
    all_papers = []
    for papers in papers_by_category.values():
        all_papers.extend(papers)

    if not all_papers:
        print("No papers to save!")
        return

    # Define field order for readability
    field_order = [
        # Identifiers
        "arxiv_id", "arxiv_id_base", "version", "s2_paper_id",
        # Core
        "title", "authors", "author_count", "author_affiliations",
        # Dates
        "published_date", "updated_date", "paper_age_days",
        # Categories
        "primary_category", "all_categories", "category_count",
        # Citations
        "citations", "influential_citations", "reference_count",
        "citations_per_month", "citations_per_year",
        # Fields & Venue
        "fields_of_study", "s2_fields_of_study", "publication_venue", "venue_type",
        "publication_types",
        # Content
        "abstract", "abstract_length", "tldr",
        "arxiv_comment", "journal_reference",
        # Access
        "is_open_access", "open_access_url",
        # Identifiers
        "doi",
        # Links
        "arxiv_url", "pdf_url", "s2_url",
        # Extra datetime fields
        "published_datetime", "updated_datetime"
    ]

    # Ensure all fields are included
    all_fields = set()
    for p in all_papers:
        all_fields.update(p.keys())

    fieldnames = [f for f in field_order if f in all_fields]
    extra_fields = sorted(all_fields - set(fieldnames))
    fieldnames.extend(extra_fields)

    # Save combined file
    combined_path = os.path.join(output_dir, "arxiv_papers_all.csv")
    with open(combined_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(all_papers)
    print(f"Saved: {combined_path} ({len(all_papers)} papers)")

    # Save per-category files
    for category, papers in papers_by_category.items():
        if not papers:
            continue

        cat_safe = category.replace(".", "_").replace("-", "_")
        cat_path = os.path.join(output_dir, f"arxiv_{cat_safe}.csv")

        with open(cat_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(papers)
        print(f"Saved: {cat_path} ({len(papers)} papers)")


def save_to_xlsx(papers_by_category: Dict[str, List[Dict]], output_dir: str):
    """Save papers to Excel with multiple sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl not installed. Skipping Excel export.")
        print("Install with: uv add openpyxl")
        return

    os.makedirs(output_dir, exist_ok=True)
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_headers = [
        "Category", "Papers", "Total Citations", "Avg Citations",
        "Influential Cites", "Avg References", "Top Venue", "Date Range"
    ]

    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row = 2
    for category, papers in sorted(papers_by_category.items()):
        if not papers:
            continue

        citations = [p["citations"] or 0 for p in papers]
        influential = [p["influential_citations"] or 0 for p in papers]
        refs = [p["reference_count"] or 0 for p in papers]
        venues = [p["publication_venue"] for p in papers if p["publication_venue"]]
        dates = [p["published_date"] for p in papers if p["published_date"]]

        # Most common venue
        venue_counts = defaultdict(int)
        for v in venues:
            venue_counts[v] += 1
        top_venue = max(venue_counts, key=venue_counts.get) if venue_counts else ""

        data = [
            category,
            len(papers),
            sum(citations),
            round(sum(citations) / len(citations), 2) if citations else 0,
            sum(influential),
            round(sum(refs) / len(refs), 2) if refs else 0,
            top_venue[:40],
            f"{min(dates)} to {max(dates)}" if dates else "N/A"
        ]

        for col, val in enumerate(data, 1):
            cell = ws_summary.cell(row=row, column=col, value=val)
            cell.alignment = cell_align
            cell.border = thin_border
        row += 1

    # Column widths for summary
    widths = [15, 10, 15, 12, 15, 12, 40, 25]
    for col, w in enumerate(widths, 1):
        ws_summary.column_dimensions[ws_summary.cell(row=1, column=col).column_letter].width = w
    ws_summary.freeze_panes = "A2"

    # Category sheets (abbreviated data for Excel)
    excel_fields = [
        "arxiv_id", "title", "authors", "published_date", "citations",
        "influential_citations", "fields_of_study", "publication_venue",
        "tldr", "arxiv_url"
    ]

    for category, papers in sorted(papers_by_category.items()):
        if not papers:
            continue

        cat_safe = category.replace(".", "_").replace("-", "_")[:31]
        ws = wb.create_sheet(title=cat_safe)

        # Headers
        for col, header in enumerate(excel_fields, 1):
            cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data
        for row_idx, paper in enumerate(papers, 2):
            for col, field in enumerate(excel_fields, 1):
                val = paper.get(field, "")
                if val is None:
                    val = ""
                # Truncate long fields for Excel
                if isinstance(val, str) and len(val) > 500:
                    val = val[:500] + "..."
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = cell_align
                cell.border = thin_border

        # Column widths
        col_widths = [15, 60, 40, 12, 10, 12, 30, 30, 60, 35]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
        ws.freeze_panes = "A2"

    xlsx_path = os.path.join(output_dir, "arxiv_papers_dataset.xlsx")
    wb.save(xlsx_path)
    print(f"Saved: {xlsx_path}")


def generate_statistics(papers_by_category: Dict[str, List[Dict]], output_dir: str):
    """Generate comprehensive statistics."""
    os.makedirs(output_dir, exist_ok=True)

    all_papers = []
    for papers in papers_by_category.values():
        all_papers.extend(papers)

    if not all_papers:
        return

    stats = {
        "generated_at": datetime.now().isoformat(),
        "total_papers": len(all_papers),
        "summary": {},
        "categories": {},
        "fields_of_study": defaultdict(int),
        "venues": defaultdict(int),
        "publication_types": defaultdict(int),
        "papers_by_month": defaultdict(int),
        "top_cited_papers": [],
        "top_influential_papers": [],
        "papers_with_tldr": 0,
        "papers_with_venue": 0,
        "papers_open_access": 0,
        "average_metrics": {}
    }

    # Calculate metrics
    all_citations = []
    all_influential = []
    all_references = []
    all_ages = []

    for paper in all_papers:
        # Citations
        if paper["citations"] is not None:
            all_citations.append(paper["citations"])
        if paper["influential_citations"] is not None:
            all_influential.append(paper["influential_citations"])
        if paper["reference_count"] is not None:
            all_references.append(paper["reference_count"])
        if paper["paper_age_days"] is not None:
            all_ages.append(paper["paper_age_days"])

        # Fields
        if paper["fields_of_study"]:
            for f in paper["fields_of_study"].split(", "):
                if f.strip():
                    stats["fields_of_study"][f.strip()] += 1

        # Venues
        if paper["publication_venue"]:
            stats["venues"][paper["publication_venue"]] += 1
            stats["papers_with_venue"] += 1

        # Publication types
        if paper["publication_types"]:
            for pt in paper["publication_types"].split(", "):
                if pt.strip():
                    stats["publication_types"][pt.strip()] += 1

        # Monthly distribution
        if paper["published_date"]:
            month = paper["published_date"][:7]
            stats["papers_by_month"][month] += 1

        # TL;DR and open access
        if paper["tldr"]:
            stats["papers_with_tldr"] += 1
        if paper["is_open_access"]:
            stats["papers_open_access"] += 1

    # Summary stats
    stats["summary"] = {
        "total_citations": sum(all_citations),
        "total_influential_citations": sum(all_influential),
        "total_references": sum(all_references),
        "papers_with_tldr": stats["papers_with_tldr"],
        "papers_with_venue": stats["papers_with_venue"],
        "papers_open_access": stats["papers_open_access"],
    }

    stats["average_metrics"] = {
        "avg_citations": round(sum(all_citations) / len(all_citations), 2) if all_citations else 0,
        "avg_influential_citations": round(sum(all_influential) / len(all_influential), 2) if all_influential else 0,
        "avg_references": round(sum(all_references) / len(all_references), 2) if all_references else 0,
        "avg_paper_age_days": round(sum(all_ages) / len(all_ages), 1) if all_ages else 0,
        "max_citations": max(all_citations) if all_citations else 0,
        "max_influential_citations": max(all_influential) if all_influential else 0,
    }

    # Per-category stats
    for category, papers in papers_by_category.items():
        if not papers:
            continue

        cat_citations = [p["citations"] or 0 for p in papers]
        cat_influential = [p["influential_citations"] or 0 for p in papers]

        stats["categories"][category] = {
            "count": len(papers),
            "total_citations": sum(cat_citations),
            "avg_citations": round(sum(cat_citations) / len(cat_citations), 2) if cat_citations else 0,
            "max_citations": max(cat_citations) if cat_citations else 0,
            "total_influential": sum(cat_influential),
        }

    # Top papers
    sorted_by_citations = sorted(
        [p for p in all_papers if p["citations"] is not None],
        key=lambda x: x["citations"],
        reverse=True
    )
    stats["top_cited_papers"] = [
        {
            "arxiv_id": p["arxiv_id"],
            "title": p["title"],
            "citations": p["citations"],
            "venue": p["publication_venue"],
            "url": p["arxiv_url"]
        }
        for p in sorted_by_citations[:30]
    ]

    sorted_by_influential = sorted(
        [p for p in all_papers if p["influential_citations"] is not None],
        key=lambda x: x["influential_citations"],
        reverse=True
    )
    stats["top_influential_papers"] = [
        {
            "arxiv_id": p["arxiv_id"],
            "title": p["title"],
            "influential_citations": p["influential_citations"],
            "citations": p["citations"],
            "url": p["arxiv_url"]
        }
        for p in sorted_by_influential[:30]
    ]

    # Convert defaultdicts to dicts for JSON
    stats["fields_of_study"] = dict(sorted(stats["fields_of_study"].items(),
                                           key=lambda x: x[1], reverse=True)[:50])
    stats["venues"] = dict(sorted(stats["venues"].items(),
                                  key=lambda x: x[1], reverse=True)[:50])
    stats["publication_types"] = dict(stats["publication_types"])
    stats["papers_by_month"] = dict(sorted(stats["papers_by_month"].items()))

    # Save statistics
    stats_path = os.path.join(output_dir, "dataset_statistics.json")
    with open(stats_path, "w", encoding="utf-8") as f:
        json.dump(stats, f, indent=2)
    print(f"Saved: {stats_path}")

    return stats


def print_summary(stats: Dict):
    """Print a summary of the collected data."""
    print("\n" + "=" * 70)
    print("DATASET SUMMARY")
    print("=" * 70)

    print(f"\nTotal papers: {stats['total_papers']:,}")
    print(f"Total citations: {stats['summary']['total_citations']:,}")
    print(f"Total influential citations: {stats['summary']['total_influential_citations']:,}")

    print(f"\nPapers with TL;DR: {stats['papers_with_tldr']:,}")
    print(f"Papers with venue: {stats['papers_with_venue']:,}")
    print(f"Open access papers: {stats['papers_open_access']:,}")

    print(f"\n--- Average Metrics ---")
    for metric, value in stats["average_metrics"].items():
        print(f"  {metric}: {value}")

    print(f"\n--- Top 10 Fields of Study ---")
    for field, count in list(stats["fields_of_study"].items())[:10]:
        print(f"  {field}: {count:,}")

    print(f"\n--- Top 10 Venues ---")
    for venue, count in list(stats["venues"].items())[:10]:
        venue_short = venue[:50] + "..." if len(venue) > 50 else venue
        print(f"  {venue_short}: {count:,}")

    print(f"\n--- Top 10 Most Cited Papers ---")
    for i, paper in enumerate(stats["top_cited_papers"][:10], 1):
        title = paper["title"][:55] + "..." if len(paper["title"]) > 55 else paper["title"]
        print(f"  {i}. [{paper['citations']:,} citations] {title}")

    print(f"\n--- Papers by Month ---")
    for month, count in stats["papers_by_month"].items():
        print(f"  {month}: {count:,}")


def main():
    args = parse_args()

    # Date range
    end_date = datetime.strptime(args.end_date, "%Y-%m-%d") if args.end_date else datetime.now()
    start_date = datetime.strptime(args.start_date, "%Y-%m-%d") if args.start_date else end_date - timedelta(days=365)

    start_str = start_date.strftime("%Y%m%d")
    end_str = end_date.strftime("%Y%m%d")

    # Categories
    if args.categories:
        categories = [c.strip() for c in args.categories.split(",")]
    elif args.all_categories:
        categories = ALL_CATEGORIES
    else:
        categories = CS_CATEGORIES

    print("=" * 70)
    print("COMPREHENSIVE ARXIV PAPER MINER")
    print("=" * 70)
    print(f"Date range: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
    print(f"Categories: {len(categories)}")
    print(f"Max results per category: {args.max_results}")
    print(f"Semantic Scholar enrichment: {'disabled' if args.skip_semantic_scholar else 'enabled'}")
    print(f"Output directory: {args.output_dir}")
    print("=" * 70)

    # Collect papers
    papers_by_category = {}

    for i, category in enumerate(categories, 1):
        print(f"\n[{i}/{len(categories)}] Processing {category}...")

        papers = fetch_arxiv_papers(
            category=category,
            start_date=start_str,
            end_date=end_str,
            max_results=args.max_results,
            rate_limit=args.rate_limit
        )

        print(f"    Found {len(papers)} papers")

        # Semantic Scholar enrichment
        if not args.skip_semantic_scholar and papers:
            print(f"    Enriching with Semantic Scholar data...")
            papers = enrich_with_semantic_scholar(papers, rate_limit=args.s2_rate_limit)

        papers_by_category[category] = papers

    # Save results
    print("\n" + "=" * 70)
    print("SAVING RESULTS")
    print("=" * 70)

    save_to_csv(papers_by_category, args.output_dir)
    save_to_xlsx(papers_by_category, args.output_dir)
    stats = generate_statistics(papers_by_category, args.output_dir)

    if stats:
        print_summary(stats)

    print("\n" + "=" * 70)
    print("DONE!")
    print("=" * 70)


if __name__ == "__main__":
    main()
